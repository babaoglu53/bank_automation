from Transactions import Ui_MainWindow
from PyQt5 import QtCore, QtWidgets
from Database import Database
import openpyxl as xlr
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class TransactionsPage(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.pushButton_show.clicked.connect(self.showDatabase)
        self.ui.pushButton_clear.clicked.connect(self.clearTable)
        self.ui.pushButton_exportToExcel.clicked.connect(self.exportToExcel)
        
        self.db = Database("localhost", 3306, "root", "", "bank")
        self.db_connection = self.db.mydb

        current_date = QtCore.QDateTime.currentDateTime()
        self.ui.dateEdit.setDateTime(current_date)
        self.ui.dateEdit.setCalendarPopup(True)
        self.ui.dateEdit_2.setDateTime(current_date)
        self.ui.dateEdit_2.setCalendarPopup(True)

        header = self.ui.tableWidget.horizontalHeader()       
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents + 10)

        self.ui.label_temp.close()
    
    def getAccountNoAndCloseLabel(self):
        account_no = self.ui.label_temp.text()
        self.ui.label_temp.close()
        return account_no

    def showDatabase(self):
        date_1 = self.ui.dateEdit.date().toString("yyyy-MM-dd 00:00:00")
        date_2 = self.ui.dateEdit_2.date().toString("yyyy-MM-dd 23:59:59")

        query_result = self.db.dateQuery("SELECT * FROM transactions WHERE account_no = {} AND (transaction_date BETWEEN '{}' AND '{}')".format(self.getAccountNoAndCloseLabel(), date_1, date_2), self.db.getCursor(self.db_connection))
    
        for row_number, row_data in zip(range(len(query_result)), query_result):
            for column_number, data in zip(range(len(row_data)), row_data[1:]):
                if str(type(data)) == "<class 'datetime.datetime'>":
                    cell_content = QtWidgets.QTableWidgetItem(data.strftime("%Y-%m-%d %H:%M:%S"))
                else:
                    cell_content = QtWidgets.QTableWidgetItem(str(data))
                self.ui.tableWidget.setItem(row_number, column_number, cell_content)

    def clearTable(self):
        row_count = self.ui.tableWidget.rowCount()

        for row_index in reversed(range(row_count)):
            self.ui.tableWidget.removeRow(row_index)

        self.ui.tableWidget.setRowCount(10)

    def exportToExcel(self):
        model = self.ui.tableWidget.model()
        data = []
        
        for row in range(model.rowCount()):
            data.append([])
            for column in range(model.columnCount()):
                index = model.index(row, column)
                data[row].append(str(model.data(index)))

        book = xlr.Workbook()
        sheet = book.active

        sheet.append(("İşlem Adı", "İşlem Tarihi", "Miktar"))
        sheet.column_dimensions['A'].width = 60
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 60

        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        for i in range(1, 4):
            sheet.cell(row = 1, column = i).font = Font(name = 'Century Gothic', bold=True, italic=True, color="00000080")
            sheet.cell(row = 1, column = i).fill = PatternFill(fill_type='solid', start_color='0000FFFF', end_color='0000FFFF')
            sheet.cell(row = 1, column = i).border = thin_border
            sheet.cell(row = 1, column = i).alignment = Alignment(horizontal='center', vertical='center')

        for i in data:
            if i != ["None", "None", "None"]:
                sheet.append(i)
                print(i)

        for i in range(1, 2):
            for j in range(2, len(sheet['A'])+1):
                sheet.cell(row = j, column = i).font = Font(name = 'Century Gothic', bold=True, italic=True, color="00000000")
                sheet.cell(row = j, column = i).fill = PatternFill(fill_type='solid', start_color='00FFFF00', end_color='00FFFF00')
                sheet.cell(row = j, column = i).border = thin_border
                sheet.cell(row = j, column = i).alignment = Alignment(horizontal='center', vertical='center')

        for i in range(2, 3):
            for j in range(2, len(sheet['B'])+1):
                sheet.cell(row = j, column = i).font = Font(name = 'Century Gothic', color="00000000")
                sheet.cell(row = j, column = i).fill = PatternFill(fill_type='solid', start_color='00FFFF00', end_color='00FFFF00')
                sheet.cell(row = j, column = i).border = thin_border

        for i in range(3, 4):
            for j in range(2, len(sheet['C'])+1):
                sheet.cell(row = j, column = i).font = Font(name = 'Century Gothic', color="00000000")
                sheet.cell(row = j, column = i).fill = PatternFill(fill_type='solid', start_color='00FFFF00', end_color='00FFFF00')
                sheet.cell(row = j, column = i).border = thin_border
                sheet.cell(row = j, column = i).alignment = Alignment(horizontal='center', vertical='center')

        book.save("hesap_gecmisi.xlsx")
        book.close()

